import tensorflow as tf
from tflearn.layers.conv import global_avg_pool
from tensorflow.contrib.layers import batch_norm, flatten
from tensorflow.contrib.layers import xavier_initializer
from tensorflow.contrib.framework import arg_scope
import cifar10_input

import pickle
import openpyxl
from xlutils.copy import copy
import xlwt
import xlrd
import os
import time
import matplotlib.pyplot as plt

from sklearn.metrics import accuracy_score, precision_score, recall_score, f1_score

# Hyperparameter
growth_k = 24
nb_block = 2  # how many (dense block + Transition Layer) ?
init_learning_rate = 0.01
epsilon = 1e-4  # AdamOptimizer epsilon
dropout_rate = 0.2

# Momentum Optimizer will use
nesterov_momentum = 0.9
weight_decay = 1e-4

# Label & batch_size
batch_size = 64

iteration = 782
# batch_size * iteration = data_set_number

test_iteration = 10

training_epochs = 200

# in cifar10
num_examples_per_epoch_for_train = cifar10_input.NUM_EXAMPLES_PER_EPOCH_FOR_TRAIN
num_examples_per_epoch_for_eval = cifar10_input.NUM_EXAMPLES_PER_EPOCH_FOR_EVAL
image_size = 32
img_channels = 3
class_num = 10
data_dir = './'



def loss_to_excel(step, cross_entropy, AdaptiveFun_name):
    path = os.path.join('./result_save/cifar10_Densenet_loss_2.xlsx')
    p = os.path.exists(path)
    loss = cross_entropy
    st = step
    if p:

        rb = openpyxl.load_workbook(path)
        wb = rb.get_sheet_names()
        # wb = copy(rb)
        s = rb.get_sheet_by_name(wb[0])
        # s = wb.get_sheet(0)

        for n in range(1, st + 1):
            # s.write(n, 0, n)
            s.cell(row=n + 1, column=1).value = n
        rb.save(path)

        # s.write(0, 0, 'step')
        s.cell(row=1, column=1).value = 'step'

        # if AdaptiveFun_name == 'SinP':
        #   col = 1

        if AdaptiveFun_name == 'Sigmoid':
            col = 1
        if AdaptiveFun_name == 'Tanh':
            col = 3
        if AdaptiveFun_name == 'Relu':
            col = 5

        if AdaptiveFun_name == 'ASigmoid':
            col = 2
        if AdaptiveFun_name == 'ATanh':
            col = 4
        if AdaptiveFun_name == 'ARelu':
            col = 6
        if AdaptiveFun_name == 'swish':
            col = 7
        if AdaptiveFun_name == 'LRelu':
            col = 8
        if AdaptiveFun_name == 'PRelu':
            col = 9

        # s.write(0, col, AdaptiveFun_name)
        s.cell(row=1, column=col + 1).value = AdaptiveFun_name

        i = 1
        while (i <= len(loss)):
            # s.write(i, col, str(loss[i - 1]))  # 像表格中写入数据
            s.cell(row=i + 1, column=col + 1).value = str(loss[i - 1])
            i += 1
        rb.save(path)


def evolution_to_excel(epoch, acc_test, pre_test, rec_test, f1_test, AdaptiveFun_name):
    path = os.path.join('./result_save/cifar10_Densenet_evolution_2.xls')
    p = os.path.exists(path)

    atest = acc_test

    ptest = pre_test

    rtest = rec_test

    ftest = f1_test

    """
    data = xlwt.Workbook(encoding='ascii')  # 创建一个workboookk
    worksheet = data.add_sheet('Sheet1')  # 添加一个Sheet工作表
    """

    if p:

        rb = xlrd.open_workbook(path)
        wb = copy(rb)
        s = wb.get_sheet(0)

        for n in range(1, epoch + 1):
            s.write(n, 0, n)
        wb.save(path)

        s.write(0, 0, 'epoches')

        # if AdaptiveFun_name == 'SinP':
        #   col = 1

        if AdaptiveFun_name == 'Sigmoid':
            col = 1
        if AdaptiveFun_name == 'Tanh':
            col = 9
        if AdaptiveFun_name == 'Relu':
            col = 17

        if AdaptiveFun_name == 'ASigmoid':
            col = 5
        if AdaptiveFun_name == 'ATanh':
            col = 13
        if AdaptiveFun_name == 'ARelu':
            col = 21
        if AdaptiveFun_name == 'swish':
            col = 25
        if AdaptiveFun_name == 'LRelu':
            col = 29
        if AdaptiveFun_name == 'PRelu':
            col = 33

        """
        if AdaptiveFun_name == 'AS_UV':
            col = 13
        if AdaptiveFun_name == 'AS_LV':
            col = 17
        if AdaptiveFun_name == 'AS_IV':
            col = 21
        if AdaptiveFun_name == 'AT_UV':
            col = 13
        if AdaptiveFun_name == 'AT_LV':
            col = 15
        if AdaptiveFun_name == 'AT_IV':
            col = 17
        if AdaptiveFun_name == 'AR_UV':
            col = 19
        if AdaptiveFun_name == 'AR_LV':
            col = 21
        if AdaptiveFun_name == 'AR_IV':
            col = 23
        """

        s.write(0, col, AdaptiveFun_name + '_acc')
        s.write(0, col + 1, AdaptiveFun_name + '_pre')
        s.write(0, col + 2, AdaptiveFun_name + '_rec')
        s.write(0, col + 3, AdaptiveFun_name + '_f1')

        i = 1
        while (i <= len(atest)):
            s.write(i, col, atest[i - 1])  # 像表格中写入数据

            s.write(i, col + 1, ptest[i - 1])  # 像表格中写入数据

            s.write(i, col + 2, rtest[i - 1])  # 像表格中写入数据

            s.write(i, col + 3, ftest[i - 1])  # 像表格中写入数据
            i += 1
        wb.save(path)


def conv_layer(input, filter, kernel, stride=1, layer_name="conv"):
    with tf.name_scope(layer_name):
        network = tf.layers.conv2d(inputs=input, use_bias=False, filters=filter, kernel_size=kernel, strides=stride,
                                   padding='SAME')
        return network


def Global_Average_Pooling(x, stride=1):
    """
    width = np.shape(x)[1]
    height = np.shape(x)[2]
    pool_size = [width, height]
    return tf.layers.average_pooling2d(inputs=x, pool_size=pool_size, strides=stride) # The stride value does not matter
    It is global average pooling without tflearn
    """

    return global_avg_pool(x, name='Global_avg_pooling')
    # But maybe you need to install h5py and curses or not


def Batch_Normalization(x, training, scope):
    with arg_scope([batch_norm],
                   scope=scope,
                   updates_collections=None,
                   decay=0.9,
                   center=True,
                   scale=True,
                   zero_debias_moving_mean=True):
        return tf.cond(training,
                       lambda: batch_norm(inputs=x, is_training=training, reuse=None),
                       lambda: batch_norm(inputs=x, is_training=training, reuse=True))


def Drop_out(x, rate, training):
    return tf.layers.dropout(inputs=x, rate=rate, training=training)


"""
def Relu(z):

    a = tf.Variable(tf.constant(1.0), name='a')
    b = tf.Variable(tf.constant(1.0), name='b')
    c = tf.Variable(tf.constant(0.0), name='c')
    d = tf.Variable(tf.constant(0.0), name='d')

    #ff = b * tf.sigmoid(a * z + c) + d  ###AS
    #ff = b * tf.tanh(a * z + c) + d  ###AT
    #ff = tf.maximum(a*z+c,b*z+d) ###AR
    ff = tf.nn.leaky_relu(z, alpha=0.2, name=None)
    #ff = z * tf.sigmoid(b * z)
    #ff = tf.nn.relu(z)
   # ff = tf.sigmoid(z)
    #ff = tf.tanh(z)
    return ff
"""
def Relu(z):
    alphas = tf.Variable(tf.constant(0.0), name='alpha')

    #alphas = tf.get_variable('alpha', z.get_shape()[-1], initializer = tf.constant_initializer(0.0), dtype=tf.float32)
    pos = tf.nn.relu(z)
    neg = alphas*(z - abs(z))*0.5
    return pos+neg

def Average_pooling(x, pool_size=[2, 2], stride=2, padding='VALID'):
    return tf.layers.average_pooling2d(inputs=x, pool_size=pool_size, strides=stride, padding=padding)


def Max_Pooling(x, pool_size=[3, 3], stride=2, padding='VALID'):
    return tf.layers.max_pooling2d(inputs=x, pool_size=pool_size, strides=stride, padding=padding)


def Concatenation(layers):
    return tf.concat(layers, axis=3)


def Linear(x):
    return tf.layers.dense(inputs=x, units=class_num, name='linear')


class DenseNet():
    def __init__(self, x, nb_blocks, filters, training):
        self.nb_blocks = nb_blocks
        self.filters = filters
        self.training = training
        self.model = self.Dense_net(x)

    def bottleneck_layer(self, x, scope):
        # print(x)
        with tf.name_scope(scope):
            x = Batch_Normalization(x, training=self.training, scope=scope + '_batch1')
            x = Relu(x)
            x = conv_layer(x, filter=4 * self.filters, kernel=[1, 1], layer_name=scope + '_conv1')
            x = Drop_out(x, rate=dropout_rate, training=self.training)

            x = Batch_Normalization(x, training=self.training, scope=scope + '_batch2')
            x = Relu(x)
            x = conv_layer(x, filter=self.filters, kernel=[3, 3], layer_name=scope + '_conv2')
            x = Drop_out(x, rate=dropout_rate, training=self.training)

            # print(x)

            return x

    def transition_layer(self, x, scope):
        with tf.name_scope(scope):
            x = Batch_Normalization(x, training=self.training, scope=scope + '_batch1')
            x = Relu(x)
            # x = conv_layer(x, filter=self.filters, kernel=[1,1], layer_name=scope+'_conv1')

            # https://github.com/taki0112/Densenet-Tensorflow/issues/10

            in_channel = x.shape[-1]

            # print(in_channel)


            x = conv_layer(x, filter=int(in_channel) * 0.5, kernel=[1, 1], layer_name=scope + '_conv1')
            x = Drop_out(x, rate=dropout_rate, training=self.training)
            x = Average_pooling(x, pool_size=[2, 2], stride=2)

            return x

    def dense_block(self, input_x, nb_layers, layer_name):
        with tf.name_scope(layer_name):
            layers_concat = list()
            layers_concat.append(input_x)

            x = self.bottleneck_layer(input_x, scope=layer_name + '_bottleN_' + str(0))

            layers_concat.append(x)

            for i in range(nb_layers - 1):
                x = Concatenation(layers_concat)
                x = self.bottleneck_layer(x, scope=layer_name + '_bottleN_' + str(i + 1))
                layers_concat.append(x)

            x = Concatenation(layers_concat)

            return x

    def Dense_net(self, input_x):
        x = conv_layer(input_x, filter=2 * self.filters, kernel=[7, 7], stride=2, layer_name='conv0')
        # x = Max_Pooling(x, pool_size=[3,3], stride=2)


        """
        for i in range(self.nb_blocks) :
            # 6 -> 12 -> 48
            x = self.dense_block(input_x=x, nb_layers=4, layer_name='dense_'+str(i))
            x = self.transition_layer(x, scope='trans_'+str(i))
        """

        x = self.dense_block(input_x=x, nb_layers=6, layer_name='dense_1')
        x = self.transition_layer(x, scope='trans_1')

        x = self.dense_block(input_x=x, nb_layers=12, layer_name='dense_2')
        x = self.transition_layer(x, scope='trans_2')

        x = self.dense_block(input_x=x, nb_layers=48, layer_name='dense_3')
        x = self.transition_layer(x, scope='trans_3')

        x = self.dense_block(input_x=x, nb_layers=32, layer_name='dense_final')

        # 100 Layer
        x = Batch_Normalization(x, training=self.training, scope='linear_batch')
        x = Relu(x)
        x = Global_Average_Pooling(x)
        x = flatten(x)
        x = Linear(x)

        # x = tf.reshape(x, [-1, 10])
        return x


def get_distorted_train_batch(data_dir, batch_size):
    data_dir = os.path.join(data_dir, 'cifar-10-batches-bin')
    images, labels = cifar10_input.distorted_inputs(data_dir=data_dir, batch_size=batch_size)
    return images, labels


def get_undistorted_eval_batch(data_dir, eval_data, batch_size):
    data_dir = os.path.join(data_dir, 'cifar-10-batches-bin')
    images, labels = cifar10_input.inputs(eval_data=eval_data, data_dir=data_dir, batch_size=batch_size)
    return images, labels


with tf.name_scope('Inputs'):
    images_holder = tf.placeholder(tf.float32, [batch_size, image_size, image_size, img_channels], name='images')
    labels_holder = tf.placeholder(tf.int32, [batch_size], name='labels')
    training_flag = tf.placeholder(tf.bool)

with tf.name_scope('Inference'):
    y_predict = DenseNet(x=images_holder, nb_blocks=nb_block, filters=growth_k, training=training_flag).model

    model_pred = tf.argmax(y_predict, 1)
    model_True = labels_holder

with tf.name_scope('Loss'):
    cross_entropy = tf.nn.sparse_softmax_cross_entropy_with_logits(labels=labels_holder, logits=y_predict)
    cross_entropy_mean = tf.reduce_mean(cross_entropy, name='xentroy_loss')
    tf.add_to_collection('losses', cross_entropy_mean)
    total_loss = tf.add_n(tf.get_collection('losses'), name='total_loss')

with tf.name_scope('Train'):
    learning_rate = tf.placeholder(tf.float32)

    optimizer = tf.train.GradientDescentOptimizer(learning_rate=learning_rate)
    global_step = tf.Variable(0, name='global_step', trainable=False, dtype=tf.int64)
    train_op = optimizer.minimize(total_loss, global_step=global_step)

with tf.name_scope('GetTrainBatch'):
    images_train, labels_train = get_distorted_train_batch(data_dir=data_dir, batch_size=batch_size)

with tf.name_scope('GetTestBatch'):
    images_test, labels_test = get_undistorted_eval_batch(eval_data=True, data_dir=data_dir, batch_size=batch_size)

init_op = tf.global_variables_initializer()

acc_test = []

pre_test = []

rec_test = []

f1_test = []

y_pred = []
y_true = []
cross_entropy = []
saver = tf.train.Saver()
# step=1
with tf.Session() as sess:
    sess.run(init_op)
    #     Run in the train
    coord = tf.train.Coordinator()
    threads = tf.train.start_queue_runners(sess=sess, coord=coord)

    total_batches = int(num_examples_per_epoch_for_train / batch_size)

    test_total_batches = int(num_examples_per_epoch_for_eval / batch_size)
    total_examples = test_total_batches * batch_size

    for epoch in range(training_epochs):
        if epoch == 50 : init_learning_rate /= 10
        if epoch == 100 : init_learning_rate /= 10
        if epoch == 250 : init_learning_rate /= 10

        for batch_idx in range(total_batches):
            images_batch, labels_batch = sess.run([images_train, labels_train])
            loss = sess.run([cross_entropy_mean, train_op],
                            feed_dict={images_holder: images_batch, labels_holder: labels_batch,
                                       learning_rate: init_learning_rate, training_flag: True})

        cross_entropy.append(loss[0])
        # print("step"+str(step),loss[0])
        # step += 1
        """
        if epoch == 399:
            saver.save(sess, "./model/relu_model_densenet_cifar10_400_3.ckpt")
        if epoch == 499:
            saver.save(sess, "./model/relu_model_densenet_cifar10_500_3.ckpt")
        """
        images_batch, labels_batch = sess.run([images_test, labels_test])
        y_pred_test, y_true_test = sess.run([model_pred, model_True],
                                            feed_dict={images_holder: images_batch, labels_holder: labels_batch,
                                                       training_flag: False})

        print('epoch', epoch)

        accuracy_test = accuracy_score(y_true_test, y_pred_test)
        print("accuracy_test", accuracy_test)
        acc_test.append(accuracy_test)



        Precision_test = precision_score(y_true_test, y_pred_test, average="weighted")
        print("Precision_test", Precision_test)
        pre_test.append(Precision_test)

        Recall_test = recall_score(y_true_test, y_pred_test, average="weighted")
        print("Recall_test", Recall_test)
        rec_test.append(Recall_test)

        f1_score_test = f1_score(y_true_test, y_pred_test, average="weighted")
        print("f1_score_test", f1_score_test)
        f1_test.append(f1_score_test)

    coord.request_stop()
    coord.join(threads)

loss_to_excel(training_epochs, cross_entropy, 'PRelu')
evolution_to_excel(training_epochs, acc_test, pre_test, rec_test, f1_test, 'PRelu')
plt.plot(cross_entropy, '-k', label='PRelu')
plt.plot(acc_test, '--k', label='accuracy')
plt.plot(pre_test, '--k', color='red', label='precision')
plt.plot(rec_test, '--k', color='green', label='recall')
plt.plot(f1_test, '--k', color='yellow', label='f1')
plt.legend()
plt.grid()
plt.show()

