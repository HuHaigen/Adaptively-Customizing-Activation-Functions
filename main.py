import argparse
import tensorflow as tf
import numpy as np
import pickle
#from CIFARHelper import CifarHelper
from LeNet_model import LeNet
from AlexNet_model import AlexNet
from VGG16_model import VGG16
from GoogLeNet_model import GoogLeNet
from ResNet_model import ResNet50

import cifar10_input

from sklearn.metrics import accuracy_score, precision_score, recall_score, f1_score
import openpyxl
from xlutils.copy import copy
import xlwt
import xlrd
import os
import time
import matplotlib.pyplot as plt


parser = argparse.ArgumentParser()

# in cifar10
num_examples_per_epoch_for_train = cifar10_input.NUM_EXAMPLES_PER_EPOCH_FOR_TRAIN
num_examples_per_epoch_for_eval = cifar10_input.NUM_EXAMPLES_PER_EPOCH_FOR_EVAL
image_size = 32
img_channels = 3
class_num = 10
data_dir = './'
batch_size = 64

iteration = 782
# batch_size * iteration = data_set_number

test_iteration = 10

training_epochs = 150

parser.add_argument("--model_type", dest='model_type', default='AlexNet', help='type of model')
parser.add_argument('--dataset_dir', dest='dataset_dir', default='cifar-10-batches-py/', help='path of the dataset')
parser.add_argument('--model_dir', dest='model_dir', default='model/', help='path to saving/restoring model')

args = parser.parse_args()


def loss_to_excel(step, cross_entropy, AdaptiveFun_name):
    path = os.path.join('./result_AlexNet/cifar10_AlexNet_loss_0.xlsx')
    p = os.path.exists(path)
    loss = cross_entropy
    st = step
    if p:

        rb = openpyxl.load_workbook(path)
        wb = rb.get_sheet_names()
        # wb = copy(rb)
        s = rb.get_sheet_by_name(wb[0])
        # s = wb.get_sheet(0)

        for n in range(1, st + 1):
            # s.write(n, 0, n)
            s.cell(row=n + 1, column=1).value = n
        rb.save(path)

        # s.write(0, 0, 'step')
        s.cell(row=1, column=1).value = 'step'

        # if AdaptiveFun_name == 'SinP':
        #   col = 1

        if AdaptiveFun_name == 'Sigmoid':
            col = 1
        if AdaptiveFun_name == 'Tanh':
            col = 3
        if AdaptiveFun_name == 'Relu':
            col = 5

        if AdaptiveFun_name == 'ASigmoid':
            col = 2
        if AdaptiveFun_name == 'ATanh':
            col = 4
        if AdaptiveFun_name == 'ARelu':
            col = 6

        # s.write(0, col, AdaptiveFun_name)
        s.cell(row=1, column=col + 1).value = AdaptiveFun_name

        i = 1
        while (i <= len(loss)):
            # s.write(i, col, str(loss[i - 1]))  # 像表格中写入数据
            s.cell(row=i + 1, column=col + 1).value = str(loss[i - 1])
            i += 1
        rb.save(path)


def evolution_to_excel(epoch, acc_test, pre_test, rec_test, f1_test, AdaptiveFun_name):
    path = os.path.join('./result/cifar10_LeNet_evolution_2.xls')
    p = os.path.exists(path)

    atest = acc_test

    ptest = pre_test

    rtest = rec_test

    ftest = f1_test

    """
    data = xlwt.Workbook(encoding='ascii')  # 创建一个workboookk
    worksheet = data.add_sheet('Sheet1')  # 添加一个Sheet工作表
    """

    if p:

        rb = xlrd.open_workbook(path)
        wb = copy(rb)
        s = wb.get_sheet(0)

        for n in range(1, epoch + 1):
            s.write(n, 0, n)
        wb.save(path)

        s.write(0, 0, 'epoches')

        # if AdaptiveFun_name == 'SinP':
        #   col = 1

        if AdaptiveFun_name == 'Sigmoid':
            col = 1
        if AdaptiveFun_name == 'Tanh':
            col = 9
        if AdaptiveFun_name == 'Relu':
            col = 17

        if AdaptiveFun_name == 'ASigmoid':
            col = 5
        if AdaptiveFun_name == 'ATanh':
            col = 13
        if AdaptiveFun_name == 'ARelu':
            col = 21

        """
        if AdaptiveFun_name == 'AS_UV':
            col = 13
        if AdaptiveFun_name == 'AS_LV':
            col = 17
        if AdaptiveFun_name == 'AS_IV':
            col = 21
        if AdaptiveFun_name == 'AT_UV':
            col = 13
        if AdaptiveFun_name == 'AT_LV':
            col = 15
        if AdaptiveFun_name == 'AT_IV':
            col = 17
        if AdaptiveFun_name == 'AR_UV':
            col = 19
        if AdaptiveFun_name == 'AR_LV':
            col = 21
        if AdaptiveFun_name == 'AR_IV':
            col = 23
        """

        s.write(0, col, AdaptiveFun_name + '_acc')
        s.write(0, col + 1, AdaptiveFun_name + '_pre')
        s.write(0, col + 2, AdaptiveFun_name + '_rec')
        s.write(0, col + 3, AdaptiveFun_name + '_f1')

        i = 1
        while (i <= len(atest)):
            s.write(i, col, atest[i - 1])  # 像表格中写入数据

            s.write(i, col + 1, ptest[i - 1])  # 像表格中写入数据

            s.write(i, col + 2, rtest[i - 1])  # 像表格中写入数据

            s.write(i, col + 3, ftest[i - 1])  # 像表格中写入数据
            i += 1
        wb.save(path)


def get_distorted_train_batch(data_dir, batch_size):
    data_dir = os.path.join(data_dir, 'cifar-10-batches-bin')
    images, labels = cifar10_input.distorted_inputs(data_dir=data_dir, batch_size=batch_size)
    return images, labels


def get_undistorted_eval_batch(data_dir, eval_data, batch_size):
    data_dir = os.path.join(data_dir, 'cifar-10-batches-bin')
    images, labels = cifar10_input.inputs(eval_data=eval_data, data_dir=data_dir, batch_size=batch_size)
    return images, labels


def main():
    # Path to loading model.
    if args.model_dir[-5:] != '.ckpt':
        MODEL_DIR = args.model_dir + args.model_type + '.ckpt'
        PRE_TRAINED = False
    else:
        # Train from pretrained model.
        MODEL_DIR = args.model_dir
        PRE_TRAINED = True
        print('Training based on {}'.format(MODEL_DIR))

    # Load data.

    with tf.name_scope('Inputs'):
        images_holder = tf.placeholder(tf.float32, [batch_size, image_size, image_size, img_channels], name='images')
        labels_holder = tf.placeholder(tf.int32, [batch_size], name='labels')
        keep_prob = tf.placeholder(tf.float32)

    if args.model_type == 'LeNet':
        model = LeNet(x=images_holder, keep_prob=keep_prob, num_classes=10)
    elif args.model_type == 'AlexNet':
        model = AlexNet(x=images_holder, keep_prob=keep_prob, num_classes=10)
    elif args.model_type == 'VGG16':
        model = VGG16(x=images_holder, keep_prob=keep_prob, num_classes=10)
    elif args.model_type == 'GoogLeNet':
        model = GoogLeNet(x=images_holder, keep_prob=keep_prob, num_classes=10)
    elif args.model_type == 'ResNet50':
        model = ResNet50(x=images_holder, keep_prob=keep_prob, num_classes=10)
    else:
        print("'{}' is not recognized. "
              "Use 'LeNet' / 'AlexNet' / 'VGG16' / 'GoogLeNet' / 'ResNet50'. ".format(args.command))

    y = model.output

    model_pred = tf.argmax(y, 1)
    model_True = labels_holder

    with tf.name_scope('Loss'):
        cross_entropy = tf.reduce_mean(tf.nn.sparse_softmax_cross_entropy_with_logits(labels=labels_holder, logits=y))
        tf.add_to_collection('losses', cross_entropy)
        total_loss = tf.add_n(tf.get_collection('losses'), name='total_loss')

    with tf.name_scope('Train'):
        learning_rate = tf.placeholder(tf.float32)

        optimizer = tf.train.GradientDescentOptimizer(learning_rate=learning_rate)
        # optimizer = tf.train.AdamOptimizer(learning_rate = 0.0001)
        global_step = tf.Variable(0, name='global_step', trainable=False, dtype=tf.int64)
        train_op = optimizer.minimize(total_loss, global_step=global_step)

    with tf.name_scope('GetTrainBatch'):
        images_train, labels_train = get_distorted_train_batch(data_dir=data_dir, batch_size=batch_size)

    with tf.name_scope('GetTestBatch'):
        images_test, labels_test = get_undistorted_eval_batch(eval_data=True, data_dir=data_dir, batch_size=batch_size)

    # Add an op to initialize the variables.
    init = tf.global_variables_initializer()

    # Add ops to save and restore all the variables.


    acc_test = []

    pre_test = []

    rec_test = []

    f1_test = []
    cross_entropy_ = []
    # init_learning_rate = 0.01


    init_learning_rate = 0.001  # 0.001

    # Add ops to save and restore all the variables.
    saver = tf.train.Saver()

    config = tf.ConfigProto()

    config.gpu_options.allow_growth = True

    with tf.Session(config=config) as sess:

        # with tf.Session() as sess:

        if PRE_TRAINED:
            # Restore variables from disk.
            saver.restore(sess, MODEL_DIR)
            print("Model restored.")
        else:
            # Initialize the variables.
            sess.run(init)
        coord = tf.train.Coordinator()
        threads = tf.train.start_queue_runners(sess=sess, coord=coord)

        total_batches = int(num_examples_per_epoch_for_train / batch_size)

        test_total_batches = int(num_examples_per_epoch_for_eval / batch_size)
        total_examples = test_total_batches * batch_size

        name_scope = tf.global_variables()
        print(name_scope)

        for epoch in range(training_epochs):
            if epoch == 70: init_learning_rate /= 10
            if epoch == 100: init_learning_rate /= 10
            # if epoch == 250 : init_learning_rate /= 10

            for batch_idx in range(total_batches):
                images_batch, labels_batch = sess.run([images_train, labels_train])
                loss = sess.run([cross_entropy, train_op],
                                feed_dict={images_holder: images_batch, labels_holder: labels_batch,
                                           learning_rate: init_learning_rate, keep_prob: 0.9})

                # output = sess.run(model_pred, feed_dict={images_holder: images_batch, labels_holder: labels_batch,
                #                            learning_rate: init_learning_rate, keep_prob: 0.9})
                # print("target", labels_batch)
                # print("output", output)
                # print(loss[0])

            cross_entropy_.append(loss[0])
            # print("step"+str(step),loss[0])
            # step += 1
            """
            if epoch == 399:
                saver.save(sess, "./model/relu_model_densenet_cifar10_400_0.ckpt")
            if epoch == 499:
                saver.save(sess, "./model/relu_model_densenet_cifar10_500_0.ckpt")
            """
            images_batch, labels_batch = sess.run([images_test, labels_test])
            y_pred_test, y_true_test = sess.run([model_pred, model_True],
                                                feed_dict={images_holder: images_batch, labels_holder: labels_batch,
                                                           keep_prob: 1.0})

            print("epoch:", epoch)
            print("loss:", loss[0])

            accuracy_test = accuracy_score(y_true_test, y_pred_test)
            print("accuracy_test", accuracy_test)
            acc_test.append(accuracy_test)

            Precision_test = precision_score(y_true_test, y_pred_test, average="weighted")
            print("Precision_test", Precision_test)
            pre_test.append(Precision_test)

            Recall_test = recall_score(y_true_test, y_pred_test, average="weighted")
            print("Recall_test", Recall_test)
            rec_test.append(Recall_test)

            f1_score_test = f1_score(y_true_test, y_pred_test, average="weighted")
            print("f1_score_test", f1_score_test)
            f1_test.append(f1_score_test)

            # save_path = saver.save(sess, MODEL_DIR)
            # print("Model saved in path: %s" % save_path)

        coord.request_stop()
        coord.join(threads)

        # Save the variables to disk.
    #    save_path = saver.save(sess, MODEL_DIR)
    # print("Model saved in path: %s" % save_path)

    loss_to_excel(training_epochs, cross_entropy_, 'ARelu')
    evolution_to_excel(training_epochs, acc_test, pre_test, rec_test, f1_test, 'ARelu')
    plt.plot(cross_entropy_, '-k', label='ARelu')
    plt.plot(acc_test, '--k', label='accuracy')
    plt.plot(pre_test, '--k', color='red', label='precision')
    plt.plot(rec_test, '--k', color='green', label='recall')
    plt.plot(f1_test, '--k', color='yellow', label='f1')
    plt.legend()
    plt.grid()
    plt.show()


if __name__ == '__main__':
    main()
